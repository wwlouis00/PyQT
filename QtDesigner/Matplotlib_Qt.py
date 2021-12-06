import sys
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QApplication, QWidget
import datetime
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
import openpyxl
from datetime import datetime, timedelta
import numpy as np
from openpyxl import load_workbook
from heapq import nsmallest


#------------------------------------------------------
#Excel設定
now_output_time = str(datetime.now().strftime('%Y-%m-%d %H-%M-%S'))+"output.xlsx" #另存新檔的檔名為現在時間
wb = openpyxl.load_workbook('../EGGI_Temperature/2021年10月27日.xlsx') #開啟2021年10月27日.xlsx資料

sh = wb['2021年10月27日']
ws = wb.active
a = sh.max_column   #a是橫
b = sh.max_row      #b是直
#------------------------------------------------------
#每個Channel的溫度陣列
CH1_data = []
CH2_data = []
CH3_data = []
CH4_data = []
CH5_data = []
CH6_data = []
CH7_data = []
CH8_data = []
CH_total = [i for i in range(1,sh.max_row + 1,1)]

ws.cell(row= 1,column=a+2,value="溫度變化" )
ws.cell(row= 1,column=a+3,value="Index")
ws.cell(row= 1,column=a+a+2,value="溫度差距")
ws.cell(row= 1,column=a+2*a+2,value="T_On Mode")
ws.cell(row= 1,column=a+3*a+2,value="T_Off Mode")

for channel_num in range(1,a-1,1):
    ws.cell(row= 1,column=a+3+channel_num,value= "CH"+str(channel_num)) #a是橫 b是直
    ws.cell(row= 1,column=a+4+channel_num+8,value= "CH"+str(channel_num))
    ws.cell(row= 1,column=a+4+channel_num+8,value= "CH"+str(channel_num))
    ws.cell(row= 1,column=a+4+channel_num+8,value= "CH"+str(channel_num))
    channel_gap = []
    channel_T_On = []

    for channel_data in range(1,b+1,1):
        num_1 = ws.cell(row = channel_data + 1,column=a+3+channel_num).value = ws.cell(row = channel_data,column= a-8+channel_num).value
        if channel_num == 1:
            CH1_data.append(num_1)
        if channel_num == 2:
            CH2_data.append(num_1)
        if channel_num == 3:
            CH3_data.append(num_1)
        if channel_num == 4:
            CH4_data.append(num_1)
        if channel_num == 5:
            CH5_data.append(num_1)
        if channel_num == 6:
            CH6_data.append(num_1)
        if channel_num == 7:
            CH7_data.append(num_1)
        if channel_num == 8:
            CH8_data.append(num_1) 
        # num_2 = ws.cell(row = channel_data + 2,column=a+3+channel_num).value = ws.cell(row = channel_data + 1,column= a-8+channel_num).value
        # data_gap = float(num_2) - float(num_1)
        # channel_gap.append(data_gap)
    # print(channel_gap)
    i = len(channel_gap) - 1
    channel_T_On = [ data_gap for data_gap in channel_gap if data_gap > 0]
    # print(channel_T_On)



class Canvas(FigureCanvas):
    def __init__(self, parent):
        fig, self.ax = plt.subplots(figsize=(8, 8), dpi=100)
        super().__init__(fig)
        self.setParent(parent)
        
        """ 
        Matplotlib Script
        """

        self.ax.plot(CH_total,CH1_data,'o-',color = 'red', label="CH1_data")    #紅
        self.ax.plot(CH_total,CH2_data,'o-',color = 'orange', label="CH2_data") #澄
        self.ax.plot(CH_total,CH3_data,'o-',color = 'yellow', label="CH3_data") #黃
        self.ax.plot(CH_total,CH4_data,'o-',color = 'green', label="CH4_data")  #綠
        self.ax.plot(CH_total,CH5_data,'o-',color = '#6F00FF', label="CH5_data")#藍
        self.ax.plot(CH_total,CH6_data,'o-',color = 'm', label="CH6_data")      #靛
        self.ax.plot(CH_total,CH7_data,'o-',color = 'purple', label="CH7_data") #紫
        self.ax.plot(CH_total,CH8_data,'o-',color = 'k', label="CH8_data")      #黑
        plt.ylim(0,130)
        # self.ax.plot(t, s,'o-',color = 'red', label="CH1_data")

        self.ax.set(xlabel='Time(s)', ylabel='Temperature',
               title='Sulink Channel Temperture')
        plt.legend(loc = "best", fontsize=10)
        self.ax.grid()

class AppDemo(QWidget)      :
    def __init__(self):
        super().__init__()
        self.resize(1000, 1000)
        self.setWindowTitle(u"速靈科溫度監控")
        chart = Canvas(self)
    

if __name__ == "__main__":
    
    app = QApplication(sys.argv)        
    demo = AppDemo()
    demo.show()
    sys.exit(app.exec_())
wb.save(now_output_time)

